"""生成 Phase 8 四类报表虚构视觉样本（甲方视觉验收对照材料）。

执行门禁：
- 只调用 build_daily_report_workbook，不修改生产代码；
- 全部虚构数据：无真实姓名/手机号/微信号/线索/商户/溯源地址；
- 生成的 xlsx 不提交 Git，仅作甲方视觉对照；
- 不启动服务/不请求 9100/不连生产库/不发送微信附件；
- 完成后 sample_alignment 仍为 NOT_VERIFIED（需甲方样例对照或书面确认才能升级）。
"""

from __future__ import annotations

import re
import sys
from datetime import date
from decimal import Decimal
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from openpyxl import load_workbook

from app.services.daily_report_excel import build_daily_report_workbook
from app.services.daily_report_service import (
    REPORT_DAILY_SALES_FEEDBACK,
    REPORT_LEAD_TRACE,
    REPORT_SALES_UNIT_COST,
    REPORT_SHORT_VIDEO_LIVE_LEAD,
    ReportBuildResult,
    ReportDiagnostic,
)

REPORT_DAY = date(2026, 7, 10)
OUT_DIR = Path(__file__).resolve().parent.parent / "docs" / "phase8-visual-samples"


def _save(result: ReportBuildResult, filename: str) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = build_daily_report_workbook(result)
    target = OUT_DIR / filename
    wb.save(target)
    print(f"  生成 {target.relative_to(OUT_DIR.parent.parent)}")


def _report1_normal_and_zero() -> ReportBuildResult:
    """报表1：短视频正常值 + 直播真实零值 + 合计汇总（数据完整，generated）。"""
    return ReportBuildResult(
        report_type=REPORT_SHORT_VIDEO_LIVE_LEAD, report_variant="default", report_day=REPORT_DAY,
        columns=("content_type", "spend_amount", "private_message_count", "retained_count",
                 "retained_rate", "visit_count", "visit_rate", "deal_count", "deal_rate"),
        rows=[
            # 短视频：正常值
            {"content_type": "短视频", "spend_amount": Decimal("1280.50"),
             "private_message_count": 120, "retained_count": 48, "retained_rate": 48 / 120,
             "visit_count": 12, "visit_rate": 12 / 48, "deal_count": 3, "deal_rate": 3 / 12},
            # 直播：真实零值（显式录入 0，非缺失）
            {"content_type": "直播", "spend_amount": Decimal("0.00"),
             "private_message_count": 0, "retained_count": 0, "retained_rate": 0.0,
             "visit_count": 0, "visit_rate": 0.0, "deal_count": 0, "deal_rate": 0.0},
            # 合计
            {"content_type": "合计", "spend_amount": Decimal("1280.50"),
             "private_message_count": 120, "retained_count": 48, "retained_rate": 48 / 120,
             "visit_count": 12, "visit_rate": 12 / 48, "deal_count": 3, "deal_rate": 3 / 12},
        ],
        diagnostics=(),
    )


def _report1_missing_partial() -> ReportBuildResult:
    """报表1：直播广告指标缺失（消耗/私信/留资率"数据源未接入"）+ 合计传播，partial。"""
    return ReportBuildResult(
        report_type=REPORT_SHORT_VIDEO_LIVE_LEAD, report_variant="default", report_day=REPORT_DAY,
        columns=("content_type", "spend_amount", "private_message_count", "retained_count",
                 "retained_rate", "visit_count", "visit_rate", "deal_count", "deal_rate"),
        rows=[
            {"content_type": "短视频", "spend_amount": Decimal("980.00"),
             "private_message_count": 90, "retained_count": 36, "retained_rate": 36 / 90,
             "visit_count": 8, "visit_rate": 8 / 36, "deal_count": 2, "deal_rate": 2 / 8},
            # 直播：广告指标缺失→消耗/私信/留资率 None（数据源未接入）；计数仍按已知
            {"content_type": "直播", "spend_amount": None,
             "private_message_count": None, "retained_count": 5, "retained_rate": None,
             "visit_count": 1, "visit_rate": 1 / 5, "deal_count": 0, "deal_rate": 0.0},
            # 合计：任一广告缺失→消耗/私信/留资率 None（数据源未接入）
            {"content_type": "合计", "spend_amount": None,
             "private_message_count": None, "retained_count": 41, "retained_rate": None,
             "visit_count": 9, "visit_rate": 9 / 41, "deal_count": 2, "deal_rate": 2 / 9},
        ],
        diagnostics=(ReportDiagnostic("live_ad_metric_missing"),),
    )


def _report2_normal() -> ReportBuildResult:
    """报表2：主工作表正常汇总 + 原始总结 3 销售结构化（generated）。"""
    return ReportBuildResult(
        report_type=REPORT_DAILY_SALES_FEEDBACK, report_variant="default", report_day=REPORT_DAY,
        columns=("paid_lead_count", "total_lead_count", "passed_count", "installment_count",
                 "full_payment_count", "showroom_car_count", "find_car_count",
                 "price_match_rate", "opening_rate", "self_feeling"),
        rows=[{
            "paid_lead_count": 28, "total_lead_count": 45, "passed_count": 19,
            "installment_count": 7, "full_payment_count": 4, "showroom_car_count": 6,
            "find_car_count": 3, "price_match_rate": 0.625, "opening_rate": 19 / 45,
            "self_feeling": "今日团队整体表现稳定，分期占比偏高，建议加强全款方案介绍与展厅车型匹配。",
        }],
        extra_sheets={"原始总结": [
            {"sales_name": "测试销售A", "overall_quality": "良好", "main_problem": "价格异议较多",
             "car_model_summary": "5系/E级", "budget_summary": "30-40万",
             "cooperation_level": "高", "today_suggestion": "跟进分期方案", "extra_feedback": "客户倾向周末试驾"},
            {"sales_name": "测试销售B", "overall_quality": "中等", "main_problem": "车型不匹配",
             "car_model_summary": "3系", "budget_summary": "20-25万",
             "cooperation_level": "中", "today_suggestion": "推荐同类车", "extra_feedback": "需找车"},
            {"sales_name": "测试销售C", "overall_quality": "优秀", "main_problem": "无",
             "car_model_summary": "E级", "budget_summary": "40万",
             "cooperation_level": "高", "today_suggestion": "促成全款", "extra_feedback": "已到店"},
        ]},
        diagnostics=(),
    )


def _report2_llm_degraded() -> ReportBuildResult:
    """报表2：LLM 降级（self_feeling 固定回退文案）+ 原始总结保留，partial。"""
    return ReportBuildResult(
        report_type=REPORT_DAILY_SALES_FEEDBACK, report_variant="default", report_day=REPORT_DAY,
        columns=("paid_lead_count", "total_lead_count", "passed_count", "installment_count",
                 "full_payment_count", "showroom_car_count", "find_car_count",
                 "price_match_rate", "opening_rate", "self_feeling"),
        rows=[{
            "paid_lead_count": 15, "total_lead_count": 30, "passed_count": 8,
            "installment_count": 3, "full_payment_count": 2, "showroom_car_count": 2,
            "find_car_count": 1, "price_match_rate": None, "opening_rate": 8 / 30,
            # LLM 降级固定文案
            "self_feeling": "摘要生成失败，原始反馈见原始总结工作表",
        }],
        extra_sheets={"原始总结": [
            {"sales_name": "测试销售A", "overall_quality": "中等", "main_problem": "预算不足",
             "car_model_summary": "3系", "budget_summary": "15万",
             "cooperation_level": "中", "today_suggestion": "推荐低配", "extra_feedback": "待二次到店"},
        ]},
        diagnostics=(
            ReportDiagnostic("daily_summary_llm_failed"),
            ReportDiagnostic("showroom_price_profile_missing"),  # 价位匹配 None（展厅未配置）
        ),
    )


def _report3_trace() -> ReportBuildResult:
    """报表3：线索（手机/微信/全部联系方式优先级）+ 销售（正常/"未分配"）+ 来源（广告ID/"未归因"）+ 意向/精准/地区，partial。"""
    return ReportBuildResult(
        report_type=REPORT_LEAD_TRACE, report_variant="created", report_day=REPORT_DAY,
        columns=("contact", "sales_name", "ad_source", "precision_status",
                 "imprecision_reason", "intention_display", "no_intention_reason",
                 "region_text", "trace_url"),
        rows=[
            # 正常行：手机号优先（同时有微信，取手机）+ 已归因 + 完整反馈；号码用脱敏格式避免碰巧命中真实用户
            {"contact": "138****0001", "sales_name": "测试销售A", "ad_source": "TEST_AD_001",
             "precision_status": "精准", "imprecision_reason": "无",
             "intention_display": "H / 5系", "no_intention_reason": "",
             "region_text": "测试城市A", "trace_url": "https://example.com/test/001"},
            # 未归因行：缺 ad_id + 缺 trace_url
            {"contact": "test_wx_B", "sales_name": "测试销售B", "ad_source": "未归因",
             "precision_status": "不精准", "imprecision_reason": "预算不匹配",
             "intention_display": "M", "no_intention_reason": "",
             "region_text": "测试城市B", "trace_url": None},
            # 未分配行：无负责人 + 无反馈
            {"contact": "139****0002", "sales_name": "未分配", "ad_source": "TEST_AD_002",
             "precision_status": None, "imprecision_reason": None,
             "intention_display": None, "no_intention_reason": None,
             "region_text": None, "trace_url": "https://example.com/test/002"},
        ],
        diagnostics=(ReportDiagnostic("trace_source_incomplete"),),  # 有未归因
    )


def _report4_unit_cost() -> ReportBuildResult:
    """报表4：销售行（成本"数据不足"）+ 未分配行 + 合计行（成本数值），含真实零值。"""
    return ReportBuildResult(
        report_type=REPORT_SALES_UNIT_COST, report_variant="default", report_day=REPORT_DAY,
        columns=("sales_name", "today_lead_count", "pass_rate", "opening_rate",
                 "total_lead_count", "total_opening_count", "total_pass_count",
                 "visit_count", "deal_count", "visit_cost", "deal_cost"),
        rows=[
            # 销售A：正常
            {"sales_name": "测试销售A", "today_lead_count": 12, "pass_rate": 8 / 12,
             "opening_rate": 9 / 12, "total_lead_count": 12, "total_opening_count": 9,
             "total_pass_count": 8, "visit_count": 4, "deal_count": 1,
             "visit_cost": None, "deal_cost": None},  # 销售级"数据不足"
            # 销售B：真实零（到店/成交 0）
            {"sales_name": "测试销售B", "today_lead_count": 6, "pass_rate": 3 / 6,
             "opening_rate": 2 / 6, "total_lead_count": 6, "total_opening_count": 2,
             "total_pass_count": 3, "visit_count": 0, "deal_count": 0,
             "visit_cost": None, "deal_cost": None},
            # 未分配行
            {"sales_name": "未分配", "today_lead_count": 3, "pass_rate": 0.0,
             "opening_rate": 0.0, "total_lead_count": 3, "total_opening_count": 0,
             "total_pass_count": 0, "visit_count": 0, "deal_count": 0,
             "visit_cost": None, "deal_cost": None},
            # 合计行：成本数值（广告有）
            {"sales_name": "合计", "today_lead_count": 21, "pass_rate": 11 / 21,
             "opening_rate": 11 / 21, "total_lead_count": 21, "total_opening_count": 11,
             "total_pass_count": 11, "visit_count": 4, "deal_count": 1,
             "visit_cost": Decimal("512.50"), "deal_cost": Decimal("2050.00")},  # 2050/4, 2050/1
        ],
        diagnostics=(),
    )


# --- 脱敏自检：生成后扫描所有样本，确认无本机路径/真实信息泄漏 ---
# Windows 路径只认反斜杠（避免 https:// 的 "s:/" 误报）
_LEAK_WIN_PATH = re.compile(r"[A-Za-z]:\\")
_LEAK_UNIX_HOME = re.compile(r"/home/|/Users/|C:\\Users")
_LEAK_LAN_IP = re.compile(r"192\.168\.|\b10\.\d+\.\d+\.\d+\b|\b172\.(1[6-9]|2\d|3[01])\.")
# 手机号加数字边界，避免浮点率值（如 0.2195121951219512）中连续数字子串误报
_LEAK_FULL_PHONE = re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)")


def _scan_leak(text: str) -> list[str]:
    """返回文本中命中的泄漏类型列表（空表示干净）。"""
    hits: list[str] = []
    if _LEAK_WIN_PATH.search(text):
        hits.append("Windows绝对路径")
    if _LEAK_UNIX_HOME.search(text):
        hits.append("本机用户路径")
    if _LEAK_LAN_IP.search(text):
        hits.append("内网IP")
    if _LEAK_FULL_PHONE.search(text):
        hits.append("完整11位手机号")
    return hits


def _verify_samples() -> None:
    """生成后自检：元数据/隐藏工作表/批注/超链接/单元格无本机路径或真实信息。

    覆盖 Excel 侧四类隐蔽载体（非仅可见单元格）：
    1. wb.properties（creator/title/description 等元数据）
    2. sheet_state != visible（隐藏工作表）
    3. cell.comment（单元格批注）
    4. cell.hyperlink.target（超链接对象，区别于文本 URL）
    以及全部单元格值（手机号脱敏为 138****0001 格式，不应出现完整 11 位）。
    """
    leaks: list[str] = []
    for f in sorted(OUT_DIR.glob("*.xlsx")):
        wb = load_workbook(f)
        props = wb.properties
        meta = (
            f"creator={props.creator!r} title={props.title!r} "
            f"description={props.description!r} subject={props.subject!r} "
            f"keywords={props.keywords!r} lastModifiedBy={props.lastModifiedBy!r}"
        )
        hidden = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]
        comments = [
            (ws.title, c.coordinate, c.comment.text)
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for c in row
            if c.comment
        ]
        hyperlinks = [
            (ws.title, c.coordinate, c.hyperlink.target)
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for c in row
            if c.hyperlink
        ]
        # 元数据层
        meta_text = meta + repr(hidden) + repr(comments) + repr(hyperlinks)
        meta_hits = _scan_leak(meta_text)
        if meta_hits:
            leaks.append(f"{f.name} 元数据/批注/超链接：{meta_hits}")
        # 单元格层
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.value is None:
                        continue
                    cell_hits = _scan_leak(str(c.value))
                    if cell_hits:
                        leaks.append(
                            f"{f.name} 单元格 {ws.title}!{c.coordinate}={c.value!r}：{cell_hits}"
                        )
    if leaks:
        print("⚠ 脱敏自检失败：")
        for lk in leaks:
            print(f"  - {lk}")
        raise SystemExit(1)
    print(f"脱敏自检通过：{len(list(OUT_DIR.glob('*.xlsx')))} 份样本无路径/真实信息泄漏")


def main() -> None:
    print(f"输出目录：{OUT_DIR}")
    print("生成四类报表虚构视觉样本（全部虚构数据，无敏感信息）：")
    _save(_report1_normal_and_zero(), "report1_留资管理_正常+真实零_虚构视觉样本.xlsx")
    _save(_report1_missing_partial(), "report1_留资管理_缺失partial_虚构视觉样本.xlsx")
    _save(_report2_normal(), "report2_销售反馈_正常+总结_虚构视觉样本.xlsx")
    _save(_report2_llm_degraded(), "report2_销售反馈_LLM降级_虚构视觉样本.xlsx")
    _save(_report3_trace(), "report3_线索溯源_正常+未归因+未分配_虚构视觉样本.xlsx")
    _save(_report4_unit_cost(), "report4_销售单车成本_销售+未分配+合计_虚构视觉样本.xlsx")
    _verify_samples()  # 生成后自检：元数据/隐藏/批注/超链接/单元格零泄漏，失败退出非零
    print("完成。sample_alignment 仍为 NOT_VERIFIED（需甲方样例对照或书面确认才能升级）。")


if __name__ == "__main__":
    main()
