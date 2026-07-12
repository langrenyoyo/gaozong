"""Phase 8 Task 6：每日销售报表 Excel writer（纯内存构建，不查数据库）。

固定入口：
- build_daily_report_workbook(result) -> Workbook：把 Task 5 的 ReportBuildResult 渲染成工作簿；
- save_workbook_version(workbook, target_path) -> (sha256, size_bytes)：原子落盘并返回指纹。

安全/口径：
- 不 import ORM/Session，不重新计算业务指标，缺失值原样来自 ReportBuildResult；
- 缺失数值单元格展示“数据源未接入/数据不足”，不伪造 0；
- 公式注入防护：销售文本/车型/备注/自由反馈去前导空白后以 = + - @ 开头的，写入前置单引号；
- 不生成任何 Excel 公式；比率 0.00%、金额 ¥#,##0.00。
"""

from __future__ import annotations

import hashlib
import logging
import os
import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.daily_report_service import (
    REPORT_DAILY_SALES_FEEDBACK,
    REPORT_LEAD_TRACE,
    REPORT_SALES_UNIT_COST,
    REPORT_SHORT_VIDEO_LIVE_LEAD,
    ReportBuildResult,
)

logger = logging.getLogger(__name__)

MISSING_TEXT = "数据源未接入"
INSUFFICIENT_COST_TEXT = "数据不足"  # 报表4 销售级成本固定文案（执行包第 194 行，禁止虚构分摊）

# 公式注入前缀（去前导空白后匹配）
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_FORMULA_CTRL_CHARS = ("\t", "\r", "\n")

# 工作表名（中文，固定）
_SHEET_NAMES = {
    REPORT_SHORT_VIDEO_LIVE_LEAD: "留资管理",
    REPORT_LEAD_TRACE: "线索溯源",
    REPORT_SALES_UNIT_COST: "销售单车成本",
}

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
# 每列固定宽度（字符数），稳定渲染
_COLUMN_WIDTH = {
    "content_type": 10,
    "lead_count": 10,
    "retained_count": 12,
    "retained_rate": 12,
    "spend_amount": 16,
    "private_message_count": 14,
    "sales_name": 14,
    "overall_quality": 14,
    "main_problem": 36,
    "car_model_summary": 28,
    "budget_summary": 20,
    "cooperation_level": 14,
    "today_suggestion": 36,
    "extra_feedback": 36,
    "lead_id": 10,
    "customer_name": 16,
    "traffic_type": 12,
    "ad_id": 18,
    "material_id": 18,
    "trace_url": 40,
    "assigned_staff_name": 14,
    "followup_content": 28,
    "assigned_count": 12,
    "visit_count": 10,
    "deal_count": 10,
    "visit_cost": 16,
    "deal_cost": 16,
    # Phase 8 合同补齐新增列
    "visit_rate": 10,
    "deal_rate": 10,
    "paid_lead_count": 12,
    "total_lead_count": 12,
    "passed_count": 12,
    "installment_count": 12,
    "full_payment_count": 14,
    "showroom_car_count": 14,
    "find_car_count": 12,
    "price_match_rate": 22,
    "opening_rate": 10,
    "pass_rate": 10,
    "total_opening_count": 12,
    "total_pass_count": 12,
    "today_lead_count": 12,
    "self_feeling": 36,
    "contact": 18,
    "ad_source": 18,
    "precision_status": 10,
    "imprecision_reason": 24,
    "intention_display": 18,
    "no_intention_reason": 24,
    "region_text": 14,
}
_DEFAULT_WIDTH = 16


def _column_format(col_name: str) -> str | None:
    """按列名推断数值格式：比率/金额/整数。"""
    if col_name.endswith("_rate"):
        return "0.00%"
    if col_name.endswith("_amount") or col_name.endswith("_cost"):
        return "¥#,##0.00"
    if col_name.endswith("_count"):
        return "#,##0"
    return None


def _sanitize_cell_value(value):
    """公式注入防护：字符串去前导空白后以 = + - @ 或控制字符开头的，前置单引号。

    openpyxl 对以 = 开头的字符串会当作公式；前置单引号使其成为纯文本。
    """
    if not isinstance(value, str):
        return value
    stripped = value.lstrip()
    if not stripped:
        return value
    if stripped[0] in _FORMULA_PREFIXES or stripped[0] in _FORMULA_CTRL_CHARS:
        return "'" + value
    return value


def _cell_value_for_output(col_name: str, raw_value, row: dict | None = None):
    """数值列 None 展示缺失文案；其余原样（经公式防护）。

    报表4 销售单车成本：销售/未分配行的到店成本/成交成本固定写'数据不足'（执行包第 194 行，
    禁止虚构分摊）；合计行 _cost None 才是'数据源未接入'（广告指标缺失）。
    """
    if raw_value is None and _column_format(col_name) is not None:
        if col_name in ("visit_cost", "deal_cost") and (row or {}).get("sales_name") != "合计":
            return INSUFFICIENT_COST_TEXT
        return MISSING_TEXT
    return _sanitize_cell_value(raw_value)


def _apply_number_format(cell, col_name: str, output_value):
    """数值单元格套用格式；缺失文案/文本不套。"""
    fmt = _column_format(col_name)
    if fmt is None or output_value == MISSING_TEXT or not isinstance(output_value, (int, float, Decimal)):
        return
    cell.number_format = fmt


def _write_sheet(wb: Workbook, *, title: str, columns: tuple[str, ...], rows: list[dict]) -> None:
    """把一张表写入工作簿：表头冻结 + 自动筛选 + 自动换行 + 稳定列宽。"""
    ws = wb.create_sheet(title=title)
    # 表头
    headers = [_header_label(col) for col in columns]
    ws.append(headers)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    # 数据行
    for row in rows:
        output_row = []
        for col in columns:
            output_row.append(_cell_value_for_output(col, row.get(col), row))
        ws.append(output_row)
    # 数值格式 + 边框 + 换行
    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.alignment = _BODY_ALIGN
            _apply_number_format(cell, col_name, cell.value)
    # 冻结表头 + 自动筛选
    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"
    # 稳定列宽
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COLUMN_WIDTH.get(col_name, _DEFAULT_WIDTH)


def _header_label(col_name: str) -> str:
    """列名转中文表头（固定映射，未覆盖的回退原列名）。"""
    labels = {
        # 报表 1 短视频/直播留资管理表
        "content_type": "来源类型",
        "spend_amount": "消耗金额",
        "private_message_count": "私信量",
        "retained_count": "留资量",
        "retained_rate": "留资率",
        "visit_count": "到店",
        "visit_rate": "到店率",
        "deal_count": "成交",
        "deal_rate": "成交率",
        # 报表 2 主工作表（每日线索销售反馈）
        "paid_lead_count": "线索数量",
        "total_lead_count": "总线索",
        "passed_count": "通过数量",
        "installment_count": "分期数量",
        "full_payment_count": "全款数量",
        "showroom_car_count": "展厅车型数量",
        "find_car_count": "找车数量",
        "price_match_rate": "价位区间与展厅价位一致比例",
        "opening_rate": "开口率",
        "self_feeling": "销售线索自我感觉",
        # 报表 2 原始总结 + 报表 3/4 共用销售字段
        "sales_name": "销售",
        "overall_quality": "整体质量",
        "main_problem": "主要问题",
        "car_model_summary": "车型情况",
        "budget_summary": "预算情况",
        "cooperation_level": "客户配合度",
        "today_suggestion": "今日建议",
        "extra_feedback": "补充反馈",
        # 报表 3 线索溯源表
        "contact": "线索",
        "ad_source": "来源",
        "precision_status": "精准",
        "imprecision_reason": "不精准原因",
        "intention_display": "意向",
        "no_intention_reason": "不意向原因",
        "region_text": "地区",
        "trace_url": "溯源",
        # 报表 4 销售单车成本表（sales_name/visit_count/deal_count/visit_cost/deal_cost 已在上面）
        "today_lead_count": "今日线索",
        "pass_rate": "通过率",
        "total_opening_count": "总开口",
        "total_pass_count": "总通过",
        "visit_cost": "到店成本",
        "deal_cost": "成交成本",
    }
    return labels.get(col_name, col_name)


def _write_feedback_extra_sheets(wb: Workbook, result: ReportBuildResult) -> None:
    """每日反馈表追加'原始总结'工作表（执行包固定 8 列结构化，不写 raw_text/parse_error）。

    LLM 摘要已在主工作表'销售线索自我感觉'单元格，不再单列'汇总'工作表。
    """
    raw_rows = result.extra_sheets.get("原始总结", [])
    _write_sheet(wb, title="原始总结", columns=(
        "sales_name", "overall_quality", "main_problem", "car_model_summary",
        "budget_summary", "cooperation_level", "today_suggestion", "extra_feedback",
    ), rows=raw_rows)


def build_daily_report_workbook(result: ReportBuildResult) -> Workbook:
    """把 ReportBuildResult 渲染成 openpyxl Workbook；不查数据库、不重算指标。"""
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认 Sheet

    if result.report_type == REPORT_DAILY_SALES_FEEDBACK:
        _write_sheet(wb, title="销售反馈", columns=result.columns, rows=result.rows)
        _write_feedback_extra_sheets(wb, result)
    else:
        title = _SHEET_NAMES.get(result.report_type, result.report_type)
        _write_sheet(wb, title=title, columns=result.columns, rows=result.rows)

    # 文档属性只写稳定元数据，不写手机号/微信号/原始请求体/token
    wb.properties.creator = "auto_wechat daily report"
    wb.properties.title = f"{result.report_type}/{result.report_variant}/{result.report_day.isoformat()}"
    wb.properties.description = f"diagnostics={len(result.diagnostics)}"
    return wb


def save_workbook_version(workbook: Workbook, target_path: Path) -> tuple[str, int]:
    """原子落盘：写临时文件 -> 计算指纹 -> rename 替换目标；返回 (sha256, size_bytes)。

    target_path 必须在受控存储目录内（由 storage 层保证）。失败时清理临时文件，不留半成品。
    """
    target_path = Path(target_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    # 临时文件与目标同目录，确保 rename 原子（同文件系统）
    fd, tmp_name = tempfile.mkstemp(
        prefix=".daily_report_", suffix=".tmp", dir=str(target_path.parent)
    )
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        workbook.save(tmp_path)  # openpyxl 写临时文件
        sha256 = _file_sha256(tmp_path)
        size = tmp_path.stat().st_size
        os.replace(tmp_path, target_path)  # 原子替换
    except Exception:
        # 失败清理临时文件，不留半成品
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise
    logger.info(
        "daily_report_excel stage=saved report_type=%s size=%s sha_prefix=%s",
        getattr(workbook.properties, "title", ""),
        size,
        sha256[:8],
    )
    return sha256, size


def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def verify_workbook_has_no_formulas(path: Path) -> bool:
    """辅助校验：工作簿不存在公式单元格（data_type='f'）。"""
    wb = load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    return False
    return True
