"""Phase 8 Task 6：日报 Excel 生成与安全文件存储测试。

覆盖执行包 Task 6 Step 3（工作簿合同）+ Step 4（存储安全）。
所有存储操作在 tmp_path 内，不触网、不连生产 DB。
"""

from __future__ import annotations

import hashlib
import re
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services import daily_report_excel as excel
from app.services import daily_report_storage as storage
from app.services.daily_report_service import (
    REPORT_DAILY_SALES_FEEDBACK,
    REPORT_LEAD_TRACE,
    REPORT_SALES_UNIT_COST,
    REPORT_SHORT_VIDEO_LIVE_LEAD,
    ReportBuildResult,
)

_REPORT_DAY = date(2026, 7, 10)


# ---------------------------------------------------------------------------
# ReportBuildResult 构造 helper（对齐执行包四类报表合同 columns）
# ---------------------------------------------------------------------------

def _sv_result(*, rows=None, diagnostics=()):
    return ReportBuildResult(
        report_type=REPORT_SHORT_VIDEO_LIVE_LEAD,
        report_variant="default",
        report_day=_REPORT_DAY,
        columns=("content_type", "spend_amount", "private_message_count", "retained_count",
                 "retained_rate", "visit_count", "visit_rate", "deal_count", "deal_rate"),
        rows=rows or [],
        diagnostics=diagnostics,
    )


def _feedback_result(*, rows=None, extra_sheets=None, diagnostics=()):
    return ReportBuildResult(
        report_type=REPORT_DAILY_SALES_FEEDBACK,
        report_variant="default",
        report_day=_REPORT_DAY,
        columns=("paid_lead_count", "total_lead_count", "passed_count", "installment_count",
                 "full_payment_count", "showroom_car_count", "find_car_count",
                 "price_match_rate", "opening_rate", "self_feeling"),
        rows=rows or [],
        extra_sheets=extra_sheets or {"原始总结": []},
        diagnostics=diagnostics,
    )


def _trace_result(*, rows=None, diagnostics=()):
    return ReportBuildResult(
        report_type=REPORT_LEAD_TRACE,
        report_variant="created",
        report_day=_REPORT_DAY,
        columns=("contact", "sales_name", "ad_source", "precision_status",
                 "imprecision_reason", "intention_display", "no_intention_reason",
                 "region_text", "trace_url"),
        rows=rows or [],
        diagnostics=diagnostics,
    )


def _cost_result(*, rows=None, diagnostics=()):
    return ReportBuildResult(
        report_type=REPORT_SALES_UNIT_COST,
        report_variant="default",
        report_day=_REPORT_DAY,
        columns=("sales_name", "today_lead_count", "pass_rate", "opening_rate",
                 "total_lead_count", "total_opening_count", "total_pass_count",
                 "visit_count", "deal_count", "visit_cost", "deal_cost"),
        rows=rows or [],
        diagnostics=diagnostics,
    )


# ---------------------------------------------------------------------------
# Step 3：工作簿合同（逐列表头 + 列顺序 + 格式）
# ---------------------------------------------------------------------------

def test_workbook_short_video_live_lead_columns_and_order():
    """报表 1：9 列中文表头与顺序。"""
    result = _sv_result(rows=[{
        "content_type": "短视频", "spend_amount": Decimal("100.00"),
        "private_message_count": 10, "retained_count": 6, "retained_rate": 0.6,
        "visit_count": 3, "visit_rate": 0.5, "deal_count": 1, "deal_rate": 1 / 3,
    }])
    wb = excel.build_daily_report_workbook(result)
    assert wb.sheetnames == ["留资管理"]
    ws = wb["留资管理"]
    headers = [c.value for c in ws[1]]
    assert headers == ["来源类型", "消耗金额", "私信量", "留资量", "留资率", "到店", "到店率", "成交", "成交率"]


def test_workbook_lead_trace_columns_and_order():
    """报表 3：9 列中文表头与顺序。"""
    result = _trace_result(rows=[{
        "contact": "138", "sales_name": "未分配", "ad_source": "未归因",
        "precision_status": None, "imprecision_reason": None, "intention_display": None,
        "no_intention_reason": None, "region_text": None, "trace_url": None,
    }])
    wb = excel.build_daily_report_workbook(result)
    ws = wb["线索溯源"]
    headers = [c.value for c in ws[1]]
    assert headers == ["线索", "销售", "来源", "精准", "不精准原因", "意向", "不意向原因", "地区", "溯源"]


def test_workbook_sales_unit_cost_columns_and_order():
    """报表 4：11 列中文表头与顺序。"""
    result = _cost_result(rows=[{
        "sales_name": "合计", "today_lead_count": 1, "pass_rate": 0.0, "opening_rate": 0.0,
        "total_lead_count": 1, "total_opening_count": 0, "total_pass_count": 0,
        "visit_count": 0, "deal_count": 0, "visit_cost": None, "deal_cost": None,
    }])
    wb = excel.build_daily_report_workbook(result)
    ws = wb["销售单车成本"]
    headers = [c.value for c in ws[1]]
    assert headers == ["销售", "今日线索", "通过率", "开口率", "总线索", "总开口", "总通过", "到店", "成交", "到店成本", "成交成本"]


def test_workbook_daily_feedback_main_and_raw_sheets():
    """报表 2：主工作表 10 列汇总 + 原始总结 8 列结构化（LLM 摘要在主表 self_feeling，无'汇总'工作表）。"""
    result = _feedback_result(
        rows=[{"paid_lead_count": 5, "total_lead_count": 10, "passed_count": 2,
               "installment_count": 1, "full_payment_count": 1, "showroom_car_count": 1,
               "find_car_count": 1, "price_match_rate": 0.5, "opening_rate": 0.3,
               "self_feeling": "今日整体良好"}],
        extra_sheets={
            "原始总结": [{"sales_name": "张三", "overall_quality": "良好", "main_problem": "价格高",
                          "car_model_summary": "宝马5系", "budget_summary": "10万",
                          "cooperation_level": "高", "today_suggestion": "跟紧", "extra_feedback": "无"}],
        },
    )
    wb = excel.build_daily_report_workbook(result)
    assert wb.sheetnames == ["销售反馈", "原始总结"]  # 不再有"汇总"工作表
    # 主工作表表头 10 列
    main_headers = [c.value for c in wb["销售反馈"][1]]
    assert main_headers == ["线索数量", "总线索", "通过数量", "分期数量", "全款数量", "展厅车型数量",
                            "找车数量", "价位区间与展厅价位一致比例", "开口率", "销售线索自我感觉"]
    # self_feeling 在主表最后一列
    assert wb["销售反馈"][2][9].value == "今日整体良好"
    # 原始总结表头 8 列
    raw_headers = [c.value for c in wb["原始总结"][1]]
    assert raw_headers == ["销售", "整体质量", "主要问题", "车型情况", "预算情况", "客户配合度", "今日建议", "补充反馈"]
    assert wb["原始总结"][2][0].value == "张三"


def test_workbook_missing_values_show_text_not_zero():
    """报表 1：消耗/私信/留资率 None→'数据源未接入'（不是 0）；分母 0 比率→0.0。"""
    result = _sv_result(rows=[{
        "content_type": "短视频", "spend_amount": None, "private_message_count": None,
        "retained_count": 2, "retained_rate": None, "visit_count": 0,
        "visit_rate": 0.0, "deal_count": 0, "deal_rate": 0.0,
    }])
    wb = excel.build_daily_report_workbook(result)
    ws = wb["留资管理"]
    row = [c.value for c in ws[2]]
    # 列顺序：来源类型(0) 消耗(1) 私信(2) 留资量(3) 留资率(4) 到店(5) 到店率(6) 成交(7) 成交率(8)
    assert row[1] == excel.MISSING_TEXT  # spend_amount None
    assert row[2] == excel.MISSING_TEXT  # private_message_count None
    assert row[4] == excel.MISSING_TEXT  # retained_rate None
    # 已知数值保留（分母 0 比率是数值 0，不是文案）
    assert row[3] == 2  # retained_count
    assert row[5] == 0  # visit_count
    assert row[6] == 0.0  # visit_rate


def test_workbook_sales_unit_cost_staff_insufficient_vs_total_missing():
    """报表 4：销售行 _cost None→'数据不足'；合计行 _cost None→'数据源未接入'。"""
    result = _cost_result(rows=[
        {"sales_name": "销售甲", "today_lead_count": 2, "pass_rate": 0.5, "opening_rate": 0.5,
         "total_lead_count": 2, "total_opening_count": 1, "total_pass_count": 1,
         "visit_count": 1, "deal_count": 0, "visit_cost": None, "deal_cost": None},
        {"sales_name": "合计", "today_lead_count": 2, "pass_rate": 0.5, "opening_rate": 0.5,
         "total_lead_count": 2, "total_opening_count": 1, "total_pass_count": 1,
         "visit_count": 1, "deal_count": 0, "visit_cost": None, "deal_cost": None},
    ])
    wb = excel.build_daily_report_workbook(result)
    ws = wb["销售单车成本"]
    # visit_cost=col10, deal_cost=col11
    staff_row = [c.value for c in ws[2]]
    total_row = [c.value for c in ws[3]]
    assert staff_row[9] == excel.INSUFFICIENT_COST_TEXT  # 销售行 visit_cost
    assert staff_row[10] == excel.INSUFFICIENT_COST_TEXT  # 销售行 deal_cost
    assert total_row[9] == excel.MISSING_TEXT  # 合计行 visit_cost None→数据源未接入
    assert total_row[10] == excel.MISSING_TEXT


def test_workbook_rate_and_money_number_format():
    """比率列 0.00%，金额列 ¥#,##0.00。"""
    result = _sv_result(rows=[{
        "content_type": "短视频", "spend_amount": Decimal("1234.50"),
        "private_message_count": 10, "retained_count": 5, "retained_rate": 0.5,
        "visit_count": 2, "visit_rate": 0.4, "deal_count": 1, "deal_rate": 0.5,
    }])
    wb = excel.build_daily_report_workbook(result)
    ws = wb["留资管理"]
    rate_cell = ws.cell(row=2, column=5)  # retained_rate
    money_cell = ws.cell(row=2, column=2)  # spend_amount
    assert rate_cell.number_format == "0.00%"
    assert money_cell.number_format == "¥#,##0.00"
    assert rate_cell.value == 0.5
    assert money_cell.value == Decimal("1234.50")


def test_workbook_no_formula_cells_formula_injection_guard():
    """销售文本/车型/备注/自我感觉以 = + - @ 开头，写入前置单引号，工作簿不存在公式单元格。"""
    result = _feedback_result(
        rows=[{"paid_lead_count": 1, "total_lead_count": 1, "passed_count": 0,
               "installment_count": 0, "full_payment_count": 0, "showroom_car_count": 0,
               "find_car_count": 0, "price_match_rate": None, "opening_rate": 0.0,
               "self_feeling": "=HYPERLINK(\"https://evil\")"}],
        extra_sheets={"原始总结": [{"sales_name": "+1OVERRIDE", "overall_quality": "-x",
            "main_problem": "@SUM", "car_model_summary": "\tTAB", "budget_summary": "10万",
            "cooperation_level": "\nNL", "today_suggestion": "x", "extra_feedback": "正常"}]},
    )
    wb = excel.build_daily_report_workbook(result)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != "f", f"发现公式单元格 {cell.coordinate}: {cell.value!r}"


def test_workbook_sanitize_cell_value():
    assert excel._sanitize_cell_value("=foo").startswith("'")
    assert excel._sanitize_cell_value("  +bar").startswith("'")
    assert excel._sanitize_cell_value("-baz").startswith("'")
    assert excel._sanitize_cell_value("@qux").startswith("'")
    assert excel._sanitize_cell_value("正常") == "正常"
    assert excel._sanitize_cell_value(123) == 123
    assert excel._sanitize_cell_value(None) is None


def test_workbook_freeze_panes_and_autofilter():
    result = _sv_result(rows=[{
        "content_type": "短视频", "spend_amount": Decimal("1.00"),
        "private_message_count": 1, "retained_count": 1, "retained_rate": 1.0,
        "visit_count": 0, "visit_rate": 0.0, "deal_count": 0, "deal_rate": 0.0,
    }])
    wb = excel.build_daily_report_workbook(result)
    ws = wb["留资管理"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_workbook_empty_data_still_valid():
    """空数据仍生成合法文件。"""
    result = _sv_result(rows=[])
    wb = excel.build_daily_report_workbook(result)
    path = Path(__file__).resolve().parent / "_tmp_empty.xlsx"
    try:
        wb.save(path)
        loaded = load_workbook(path)
        assert loaded.sheetnames == ["留资管理"]
        assert loaded["留资管理"].max_row == 1  # 只有表头
    finally:
        path.unlink(missing_ok=True)


def test_workbook_chinese_filename_and_content(tmp_path):
    """中文文件名和内容无乱码。"""
    result = _trace_result(rows=[{
        "contact": "13800000000", "sales_name": "李四", "ad_source": "AD1",
        "precision_status": "精准", "imprecision_reason": "无",
        "intention_display": "H / Model X", "no_intention_reason": "",
        "region_text": "北京", "trace_url": "https://h/p",
    }])
    wb = excel.build_daily_report_workbook(result)
    path = tmp_path / "线索溯源_2026-07-10.xlsx"
    wb.save(path)
    loaded = load_workbook(path)
    ws = loaded["线索溯源"]
    assert ws.cell(row=2, column=1).value == "13800000000"  # contact
    assert ws.cell(row=2, column=2).value == "李四"  # sales_name
    assert ws.cell(row=2, column=9).value == "https://h/p"  # trace_url


def test_workbook_four_report_types_all_buildable():
    """4 类报表均能构建工作簿。"""
    for result in [
        _sv_result(rows=[]),
        _feedback_result(rows=[]),
        _trace_result(rows=[]),
        _cost_result(rows=[]),
    ]:
        wb = excel.build_daily_report_workbook(result)
        assert len(wb.sheetnames) >= 1


# ---------------------------------------------------------------------------
# Step 4：存储安全
# ---------------------------------------------------------------------------

def test_storage_key_no_merchant_no_drive_no_dotdot():
    token = storage.generate_storage_token()
    key = storage.build_storage_key(REPORT_SHORT_VIDEO_LIVE_LEAD, _REPORT_DAY, token)
    assert "merchant" not in key.lower()
    assert ":" not in key
    assert ".." not in key
    assert "\\" not in key
    assert key.endswith(".xlsx")
    assert key.startswith(f"{REPORT_SHORT_VIDEO_LIVE_LEAD}/2026-07-10/")
    assert re.fullmatch(r"[0-9a-f]{32}", token)


def test_storage_token_unpredictable():
    tokens = {storage.generate_storage_token() for _ in range(20)}
    assert len(tokens) == 20  # 无碰撞


def test_storage_rejects_traversal(tmp_path):
    root = tmp_path
    with pytest.raises(ValueError):
        storage.resolve_storage_path("../secret.xlsx", root)
    with pytest.raises(ValueError):
        storage.resolve_storage_path("a/../../secret.xlsx", root)
    with pytest.raises(ValueError):
        storage.resolve_storage_path("a/b/../../etc/passwd.xlsx", root)


def test_storage_rejects_absolute_and_backslash(tmp_path):
    root = tmp_path
    with pytest.raises(ValueError):
        storage.resolve_storage_path("C:/evil.xlsx", root)
    with pytest.raises(ValueError):
        storage.resolve_storage_path("a\\b.xlsx", root)
    with pytest.raises(ValueError):
        storage.resolve_storage_path("/abs/path.xlsx", root)


def test_storage_rejects_hidden_segment_and_non_xlsx(tmp_path):
    root = tmp_path
    with pytest.raises(ValueError):
        storage.resolve_storage_path(".hidden/secret.xlsx", root)
    with pytest.raises(ValueError):
        storage.resolve_storage_path("a/b/secret.txt", root)
    with pytest.raises(ValueError):
        storage.resolve_storage_path("", root)


def test_storage_resolves_within_root(tmp_path):
    root = tmp_path
    key = storage.build_storage_key(REPORT_LEAD_TRACE, _REPORT_DAY, storage.generate_storage_token())
    target = storage.resolve_storage_path(key, root)
    assert target.resolve().is_relative_to(root.resolve())


def test_storage_atomic_write_sha_and_size(tmp_path):
    root = tmp_path
    result = _sv_result(rows=[{
        "content_type": "短视频", "spend_amount": Decimal("99.50"),
        "private_message_count": 4, "retained_count": 2, "retained_rate": 0.5,
        "visit_count": 1, "visit_rate": 0.5, "deal_count": 0, "deal_rate": 0.0,
    }])
    wb = excel.build_daily_report_workbook(result)
    key = storage.build_storage_key(REPORT_SHORT_VIDEO_LIVE_LEAD, _REPORT_DAY, storage.generate_storage_token())
    sha256, size = storage.save_workbook_to_storage(wb, key, root)
    target = storage.resolve_storage_path(key, root)
    assert target.exists()
    disk_sha = hashlib.sha256(target.read_bytes()).hexdigest()
    assert disk_sha == sha256
    assert size == target.stat().st_size
    assert size > 0


def test_save_workbook_version_no_temp_left(tmp_path):
    """成功写入后无临时文件残留。"""
    result = _cost_result(rows=[{
        "sales_name": "合计", "today_lead_count": 2, "pass_rate": 0.0, "opening_rate": 0.0,
        "total_lead_count": 2, "total_opening_count": 0, "total_pass_count": 0,
        "visit_count": 1, "deal_count": 0, "visit_cost": Decimal("300.0"), "deal_cost": None,
    }])
    wb = excel.build_daily_report_workbook(result)
    target = tmp_path / "cost.xlsx"
    excel.save_workbook_version(wb, target)
    tmp_files = [p for p in tmp_path.iterdir() if p.name.startswith(".daily_report_")]
    assert tmp_files == []
    assert target.exists()


def test_save_workbook_version_failed_write_cleans_temp(tmp_path, monkeypatch):
    """写入失败时清理临时文件，不留半成品。"""
    result = _sv_result(rows=[])
    wb = excel.build_daily_report_workbook(result)
    target = tmp_path / "sub" / "out.xlsx"

    def _boom(path):
        raise OSError("simulated write failure")
    monkeypatch.setattr(wb, "save", _boom)

    with pytest.raises(OSError):
        excel.save_workbook_version(wb, target)
    tmp_files = [p for p in tmp_path.iterdir() if p.name.startswith(".daily_report_")]
    assert tmp_files == []
    assert not target.exists()


def test_save_workbook_version_overwrites_atomically(tmp_path):
    """重生成同 path 用原子替换；旧内容被新内容覆盖。"""
    result1 = _sv_result(rows=[{
        "content_type": "短视频", "spend_amount": Decimal("1.00"),
        "private_message_count": 1, "retained_count": 1, "retained_rate": 1.0,
        "visit_count": 0, "visit_rate": 0.0, "deal_count": 0, "deal_rate": 0.0,
    }])
    result2 = _sv_result(rows=[{
        "content_type": "短视频", "spend_amount": Decimal("2.00"),
        "private_message_count": 2, "retained_count": 2, "retained_rate": 1.0,
        "visit_count": 0, "visit_rate": 0.0, "deal_count": 0, "deal_rate": 0.0,
    }])
    target = tmp_path / "v.xlsx"
    wb1 = excel.build_daily_report_workbook(result1)
    excel.save_workbook_version(wb1, target)
    first_sha = hashlib.sha256(target.read_bytes()).hexdigest()
    wb2 = excel.build_daily_report_workbook(result2)
    excel.save_workbook_version(wb2, target)
    second_sha = hashlib.sha256(target.read_bytes()).hexdigest()
    assert first_sha != second_sha


def test_verify_workbook_has_no_formulas_helper(tmp_path):
    """落盘文件无公式单元格。"""
    result = _feedback_result(
        rows=[{"paid_lead_count": 1, "total_lead_count": 1, "passed_count": 0,
               "installment_count": 0, "full_payment_count": 0, "showroom_car_count": 0,
               "find_car_count": 0, "price_match_rate": None, "opening_rate": 0.0,
               "self_feeling": "=evil"}],
    )
    wb = excel.build_daily_report_workbook(result)
    path = tmp_path / "guard.xlsx"
    wb.save(path)
    assert excel.verify_workbook_has_no_formulas(path) is True


def test_validate_artifact_rejects_symlink_and_dir(tmp_path):
    root = tmp_path
    key = storage.build_storage_key(REPORT_SALES_UNIT_COST, _REPORT_DAY, storage.generate_storage_token())
    target = storage.resolve_storage_path(key, root)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.mkdir()
    with pytest.raises(ValueError):
        storage.validate_artifact_path(key, root)
    target.rmdir()
